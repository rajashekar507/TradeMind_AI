"""
TradeMind_AI: Excel Report Generator
Generates comprehensive trading reports in Excel format
"""

import pandas as pd
import os
from datetime import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelReporter:
    def __init__(self):
        """Initialize Excel Reporter"""
        print("📊 Initializing Excel Report Generator...")
        self.reports_dir = "reports"
        
        # Create reports directory if it doesn't exist
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
            
    def generate_daily_report(self):
        """Generate comprehensive daily trading report"""
        try:
            # Load trades data
            trades_file = os.path.join("data", "trades_database.json")
            if os.path.exists(trades_file):
                with open(trades_file, 'r') as f:
                    trades_data = json.load(f)
            else:
                trades_data = {"trades": [], "summary": {}}
            
            # Create workbook
            wb = Workbook()
            
            # Sheet 1: Trade Summary
            ws1 = wb.active
            ws1.title = "Trade Summary"
            self._create_summary_sheet(ws1, trades_data)
            
            # Sheet 2: Detailed Trades
            ws2 = wb.create_sheet("Detailed Trades")
            self._create_trades_sheet(ws2, trades_data)
            
            # Sheet 3: Performance Analytics
            ws3 = wb.create_sheet("Performance Analytics")
            self._create_analytics_sheet(ws3, trades_data)
            
            # Sheet 4: Risk Analysis
            ws4 = wb.create_sheet("Risk Analysis")
            self._create_risk_sheet(ws4, trades_data)
            
            # Save file
            filename = f"TradeMind_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            wb.save(filepath)
            
            print(f"✅ Excel report generated: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Error generating Excel report: {e}")
            return None
    
    def _create_summary_sheet(self, ws, trades_data):
        """Create summary sheet with key metrics"""
        # Title
        ws['A1'] = 'TRADEMIND AI - TRADING SUMMARY'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        
        # Date
        ws['A3'] = 'Report Date:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Summary metrics
        summary = trades_data.get('summary', {})
        
        ws['A5'] = 'PERFORMANCE METRICS'
        ws['A5'].font = Font(size=12, bold=True)
        ws['A5'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws['A5'].font = Font(color="FFFFFF", bold=True)
        
        metrics = [
            ('Total Trades', summary.get('total_trades', 0)),
            ('Winning Trades', summary.get('winning_trades', 0)),
            ('Losing Trades', summary.get('losing_trades', 0)),
            ('Total P&L', f"₹{summary.get('total_pnl', 0):,.2f}"),
            ('Best Trade', f"₹{summary.get('best_trade', 0):,.2f}"),
            ('Worst Trade', f"₹{summary.get('worst_trade', 0):,.2f}"),
            ('Win Rate', f"{(summary.get('winning_trades', 0) / max(summary.get('total_trades', 1), 1) * 100):.1f}%")
        ]
        
        row = 7
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
            
        # Style adjustments
        for row in ws.iter_rows(min_row=7, max_row=13, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
                if cell.column == 1:
                    cell.font = Font(bold=True)
    
    def _create_trades_sheet(self, ws, trades_data):
        """Create detailed trades sheet"""
        trades = trades_data.get('trades', [])
        
        if trades:
            # Convert to DataFrame
            df = pd.DataFrame(trades)
            
            # Select important columns
            columns = ['trade_id', 'timestamp', 'symbol', 'option_type', 'action', 
                      'quantity', 'entry_price', 'exit_price', 'pnl', 'status']
            
            # Filter columns that exist
            available_columns = [col for col in columns if col in df.columns]
            df_filtered = df[available_columns]
            
            # Write to sheet
            for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    
            # Style header row
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                
    def _create_analytics_sheet(self, ws, trades_data):
        """Create performance analytics sheet"""
        ws['A1'] = 'PERFORMANCE ANALYTICS'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Add analytics data
        ws['A3'] = 'Time-based Analysis'
        ws['A4'] = 'Hour'
        ws['B4'] = 'Trades'
        ws['C4'] = 'Win Rate'
        ws['D4'] = 'Avg P&L'
        
        # Style headers
        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            
    def _create_risk_sheet(self, ws, trades_data):
        """Create risk analysis sheet"""
        ws['A1'] = 'RISK ANALYSIS'
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A3'] = 'Risk Metrics'
        ws['A4'] = 'Metric'
        ws['B4'] = 'Value'
        
        risk_metrics = [
            ('Max Drawdown', '₹2,000'),
            ('Risk per Trade', '1%'),
            ('Max Daily Loss', '₹3,000'),
            ('Current Exposure', '₹5,000'),
            ('Risk-Reward Ratio', '1:2')
        ]
        
        row = 5
        for metric, value in risk_metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1

# Test function
if __name__ == "__main__":
    reporter = ExcelReporter()
    report_path = reporter.generate_daily_report()
    if report_path:
        print(f"📊 Report saved at: {report_path}")
        